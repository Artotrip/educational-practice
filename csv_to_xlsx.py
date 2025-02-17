import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Пути к файлам
csv_file = "flights_data_yandex.csv"
output_file = "flights_data_yandex_xl.xlsx"

# Загружаем CSV в DataFrame
df = pd.read_csv(csv_file)

# Добавляем новые столбцы с пустыми значениями
df["Дата внесения"] = ""
df["Тип воздушного судна"] = ""
df["Кол-во перевезенных пассажиров"] = ""
df["Кол-во перевезенных детей до 12 лет"] = ""

# Создаем новый Excel-файл
wb = Workbook()
ws = wb.active

# Записываем заголовки с полужирным выделением и без окраски (если столбец не окрашен)
bold_font = Font(bold=True)
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                     bottom=Side(style='thin'))

# Окрашивание столбцов
red_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

# Ищем нужные столбцы
min_price_col = None
dep_time_col = None
arr_time_col = None
for col_num, column_title in enumerate(df.columns, start=1):
    if "минимальная цена" in column_title.lower():
        min_price_col = col_num
    if "время вылета" in column_title.lower():
        dep_time_col = col_num
    if "время прилета" in column_title.lower():
        arr_time_col = col_num

# Применяем стили к заголовкам
for col_num, column_title in enumerate(df.columns, start=1):
    cell = ws.cell(row=1, column=col_num, value=column_title)
    cell.font = bold_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border
    ws.row_dimensions[1].height = 70  # Увеличенная высота заголовка

    if col_num == min_price_col:
        cell.fill = red_fill
    elif col_num == dep_time_col or col_num == arr_time_col:
        cell.fill = green_fill

# Записываем данные
for row_num, row in enumerate(df.itertuples(index=False), start=2):
    for col_num, value in enumerate(row, start=1):
        cell = ws.cell(row=row_num, column=col_num, value=value)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

        if col_num == min_price_col:
            cell.fill = red_fill
            cell.font = Font(bold=True)
        elif col_num == dep_time_col or col_num == arr_time_col:
            cell.fill = green_fill

# Автоматическая ширина столбцов
for col_num, column_title in enumerate(df.columns, start=1):
    max_length = max(df[column_title].astype(str).map(len).max(), len(column_title)) + 4
    ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = max_length

# Включаем фильтры
ws.auto_filter.ref = ws.dimensions

# Сохраняем новый файл
wb.save(output_file)
print(f"Файл сохранён: {output_file}")